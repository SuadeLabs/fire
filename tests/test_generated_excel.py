import unittest

from openpyxl.utils import get_column_letter

from fireschematoexcel.schematoexcel import read_json, test_jsontoexcel

from fireschematoexcel.schematoexcel import JSON_SCHEMAS, PRODUCT_SEED

from . import schema_enum_registry


class TestGeneratedExcel(unittest.TestCase):
    wb = test_jsontoexcel()

    def _get_tbl_hdrs(self, ws, col, row):
        hdrs = {}
        while ws[f"{get_column_letter(col)}{row}"].value is not None:
            hdrs[ws[f"{get_column_letter(col)}{row}"].value] = col
            col += 1
        return hdrs

    def _get_enum_name(self, enum, tbl_hdrs):
        name = None
        for hdr in tbl_hdrs:
            if enum in hdr:
                name = hdr
                break
        return name

    def test_schema_excel_properties(self):
        for schema in JSON_SCHEMAS:
            ws = self.wb[schema]
            js_fl = read_json(JSON_SCHEMAS[schema], PRODUCT_SEED)
            # get top left cell of excel table
            # excluding metadata
            row = 1
            col = 2
            while 1:
                if ws[f"B{row}"].value == "property" or row == 100:
                    break
                row += 1

            # test that the table exists
            self.assertEqual(
                ws[f"B{row}"].value, "property"
            )

            tbl_hdrs = self._get_tbl_hdrs(ws, col, row)
            row += 1

            for prop in js_fl["properties"]:
                # test property name
                self.assertEqual(
                    prop, ws[f'{get_column_letter(tbl_hdrs["property"])}{row}'].value # noqa
                )

                for atr in js_fl["properties"][prop]:
                    # test property attribute names
                    self.assertTrue(
                        atr in tbl_hdrs
                    )
                    if atr not in ["enum", "items"]:
                        # test property attribute values
                        self.assertEqual(
                            js_fl["properties"][prop][atr], ws[f'{get_column_letter(tbl_hdrs[atr])}{row}'].value # noqa
                        )
                row += 1

    def test_excel_enums(self):
        for schema in JSON_SCHEMAS:
            js_enum = schema_enum_registry(schema)
            if js_enum == {}:
                continue
            ws_enum = self.wb[f"{schema}_enums"]

            # get top left cell of excel table
            row = 1
            col = 2
            while 1:
                if ws_enum[f"B{row}"].value is not None or row == 100:
                    break
                row += 1

            # test that the table exists
            self.assertTrue(
                ws_enum[f"B{row}"].value is not None
            )

            tbl_hdrs = self._get_tbl_hdrs(ws_enum, col, row)

            row += 1
            tbl_start_row = row
            for enum in js_enum:
                row = tbl_start_row
                for atr in js_enum[enum]:
                    # Enums which are imported from the common json
                    # will use the schema property name in the excel
                    # resulting in divergent names i.e.:
                    # base_currency_code vs currency_code
                    if enum in tbl_hdrs:
                        # test property attribute values
                        self.assertEqual(
                            atr, ws_enum[f'{get_column_letter(tbl_hdrs[enum])}{row}'].value # noqa
                        )
                        row += 1
                    else:
                        proxy = self._get_enum_name(enum, tbl_hdrs)
                        # test property attribute values
                        self.assertEqual(
                            atr, ws_enum[f'{get_column_letter(tbl_hdrs[proxy])}{row}'].value # noqa
                        )
                        row += 1
