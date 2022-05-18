import os
import json

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# Error for unhandled exceptions
class UnhandledException(Exception):
    """Raised when an unhandled Exception occurs"""
    def __init__(self, message="An unhandled exception occured!"):
        self.message = message
        super().__init__(self.message)


# Styles
# Monospaced typeface required in order to
# assign column width based on cell content
content_font = Font(
    name='DejaVu Sans Mono',
    size=10,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000'
)

header_font = Font(
    name='DejaVu Sans Mono',
    size=10,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000'
)

# header allignment
header_alignment = Alignment(
    horizontal='center',
    vertical='center'
)

# colours
enumFill = PatternFill(
    start_color='44ad70',
    end_color='44ad70',
    fill_type='solid'
)

blankFill = PatternFill(
    start_color='757a82',
    end_color='757a82',
    fill_type='solid'
)

# Cell borders
content_border = Border(
    left=Side(
        border_style='medium',
        color='FF000000'
    ),
    right=Side(
        border_style='medium',
        color='FF000000'),
    top=Side(
        border_style='thin',
        color='FF000000'),
    bottom=Side(
        border_style='thin',
        color='FF000000'
    )
)

header_border = Border(
    left=Side(
        border_style='medium',
        color='FF000000'),
    right=Side(
        border_style='medium',
        color='FF000000'),
    top=Side(
        border_style='medium',
        color='FF000000'),
    bottom=Side(
        border_style='medium',
        color='FF000000'
    )
)

# directory paths
dir_path = os.path.dirname(os.path.realpath(__file__))

base_path = dir_path.replace(os.sep + dir_path.split(os.sep)[-1], "")
excel_file = os.path.join(dir_path, "fire_schemas.xlsx")

# constants
SCHEMA_TAB_COLOR = "5ca0db"
ENUM_TAB_COLOR = "5cdb9a"
ITEM_TAB_COLOR = "dbc05c"

SCHEMAS_DIR = os.path.join(base_path, "v1-dev")
_, _, filenames = next(os.walk(SCHEMAS_DIR), (None, None, []))
SCHEMA_FILES = [f for f in filenames if f.endswith(".json") and "common" not in f and "entity" not in f] # noqa

SCHEMA_NAMES = [n.replace(".json", "") for n in SCHEMA_FILES]

PRODUCT_SEED = os.path.join(base_path, "v1-dev", "common.json")
ENTITY_SEED = os.path.join(base_path, "v1-dev", "entity.json")

JSON_SCHEMAS = dict([(f.split(".json")[0], os.path.join(base_path, "v1-dev", f)) for f in SCHEMA_FILES]) # noqa


def get_ws_headers(data):
    """
    returns the schema metadata
    (all attributes excluding the $ref and properties)
    """
    headers = {}
    for key in data:
        if key[0] != "$" and key != "properties" and key != "allOf":
            headers[key] = data[key]
    return headers


def get_prop_columns(data):
    """
    returns all property attributes found in the json schema
    """
    columns = []
    for prop in data["properties"]:
        for key in data["properties"][prop]:
            columns.append(key)
    columns = list(set(columns))
    # sanitize columns
    for col in columns:
        if col[0] == "$":
            columns.remove(col)
        # move description to the end since it's plain text
        # and sometimes quite large
        if "description" in columns:
            columns.remove("description")
            columns.append("description")
    return columns


def write_excel(wb, ws_name, schema_data, rootschema=False, ws_row=False, ws_col=False): # noqa
    """
    function used to write schema excel sheet
    """
    if ws_name not in wb.sheetnames:
        ws_headers = get_ws_headers(schema_data)

        # reset global enum column
        cur_enum_col = 2

        # create sheet for schema
        ws = wb.create_sheet(ws_name)

        # log required attributes
        required_atrs = []
        if "required" in ws_headers:
            required_atrs = ws_headers["required"]

        # extract requirements from adjustment schema
        if "oneOf" in ws_headers:
            for itm in ws_headers["oneOf"]:
                key = [k for k in itm.keys()][0]
                for rq in itm[key]:
                    required_atrs.append(rq)
            required_atrs = list(set(required_atrs))

        # write schema metadata
        row = write_ws_headers(ws, ws_headers)
        row += 2
        hdr_col_row = row
        col = 2

        # write column headers for <schema> properties
        if "properties" in schema_data:
            prop_columns = get_prop_columns(schema_data)
            properties = schema_data["properties"]
            ws.cell(row=hdr_col_row, column=col).value = "property"
            ws.cell(row=hdr_col_row, column=col).font = header_font
            ws.cell(row=hdr_col_row, column=col).alignment = header_alignment
            ws.cell(row=hdr_col_row, column=col).border = header_border
            hdr_cols = {"property": col}
            for hdr in prop_columns:
                col += 1
                ws.cell(row=hdr_col_row, column=col).value = hdr
                ws.cell(row=hdr_col_row, column=col).font = header_font
                ws.cell(row=hdr_col_row, column=col).alignment = header_alignment # noqa
                ws.cell(row=hdr_col_row, column=col).border = header_border
                hdr_cols[hdr] = col
            row += 1

            # write property row for each property in the <schema>
            for prop in properties:
                row, cur_enum_col = json_prop_to_excel_row(wb, ws_name, ws, prop, properties[prop], hdr_cols, row, cur_enum_col) # noqa
                if prop in required_atrs:
                    ws.cell(row=row-1, column=2).font = header_font

    if rootschema:
        # add link inside root schema
        # used for attributes which contain items,
        # since these items can be viewed as pseudoschemas
        # and are created as a separate worksheet for readability
        ws = wb[rootschema]
        ws.cell(row=ws_row, column=ws_col).value = f'=HYPERLINK("#{ws_name}!B2", "{ws_name}")' # noqa
        ws.cell(row=ws_row, column=ws_col).font = content_font
        ws.cell(row=ws_row, column=ws_col).border = content_border


def get_enum_ws(wb, enum_name):
    if f"{enum_name}_enums" not in wb.sheetnames:
        enum_ws = wb.create_sheet(f"{enum_name}_enums")
    else:
        enum_ws = wb[f"{enum_name}_enums"]
    return enum_ws


def create_enum(wb, ws_name, row, col, name, enum_vals, cur_enum_col):
    """
    creates enum table inside the <schema>_enums sheet
    and adds data validation drop down list inside the
    appropriate cell of the <schema> sheet
    """
    ws = wb[ws_name]
    enum_ws = get_enum_ws(wb, ws_name)

    # add header to enum column
    enum_ws[f"{get_column_letter(cur_enum_col)}2"].value = name
    enum_ws[f"{get_column_letter(cur_enum_col)}2"].font = header_font
    enum_ws[f"{get_column_letter(cur_enum_col)}2"].alignment = header_alignment
    enum_ws[f"{get_column_letter(cur_enum_col)}2"].border = header_border

    # add enum data to column
    for i in range(len(enum_vals)):
        enum_ws[f"{get_column_letter(cur_enum_col)}{i+3}"].value = enum_vals[i]
        enum_ws[f"{get_column_letter(cur_enum_col)}{i+3}"].font = content_font
        enum_ws[f"{get_column_letter(cur_enum_col)}{i+3}"].border = content_border # noqa

    # create data validation formula based on:
    # sheet name, and collumn range of the enum data
    # and apply it to the appropriate cell inside
    # the <schema> sheet
    data_val = DataValidation(
        type="list",
        formula1=f"={enum_ws._WorkbookChild__title}!${get_column_letter(cur_enum_col)}$3:${get_column_letter(cur_enum_col)}${len(enum_vals)+2}") # noqa
    ws.add_data_validation(data_val)
    data_val.add(ws[f"{get_column_letter(col)}{row}"])

    # colour the enum drop down list green in the <schema> sheet
    ws[f"{get_column_letter(col)}{row}"].fill = enumFill
    ws[f"{get_column_letter(col)}{row}"].font = content_font
    ws[f"{get_column_letter(col)}{row}"].border = content_border

    # increment enum column
    return cur_enum_col + 1


def json_prop_to_excel_row(wb, ws_name, ws, prop, properties, hdr_cols, row, cur_enum_col): # noqa
    """
    converts the property and associated attributes
    from the json dictionary to an excel row:
    - enums are converted into a dropdown list
    - items are converted into separate pseudoschema worksheets
    linked to via hyperlink
    this ensures human readability and an organised way
    of displaying property information
    """
    ws.cell(row=row, column=hdr_cols["property"]).value = prop
    ws.cell(row=row, column=hdr_cols["property"]).font = content_font
    ws.cell(row=row, column=hdr_cols["property"]).border = content_border
    for key in properties:
        try:
            ws.cell(row=row, column=hdr_cols[key]).value = properties[key]
            ws.cell(row=row, column=hdr_cols[key]).font = content_font
            ws.cell(row=row, column=hdr_cols[key]).border = content_border
        except ValueError:
            if key == "enum":
                cur_enum_col = create_enum(wb, ws_name, row, hdr_cols[key], prop, properties[key], cur_enum_col) # noqa
            elif key == "items":
                write_excel(wb, prop, properties[key], ws_name, row, hdr_cols[key]) # noqa
            else:
                raise UnhandledException  # stop execution in case of unhandled exception # noqa

    return row + 1, cur_enum_col


def get_list_items(ref_list):
    """
    return a dict of allowed data items
    used for the batch.json schema
    """
    schemas = "["
    for itm in ref_list:
        key = [k for k in itm.keys()][0]
        schemas += f'{itm[key].split("/")[-1].replace(".json#", "")}|'
    return schemas[:-1] + "]"


def get_required_items(req_list):
    """
    convert required properties list
    to string
    """
    reqs = "["
    for rq in req_list:
        reqs += f'{rq}|'
    return reqs[:-1] + "]"


def get_contribution_requirements(req_dict):
    """
    load the contribution.json schema requirements
    as a string
    """
    reqs = "["
    for itm in req_dict:
        key = [k for k in itm.keys()][0]
        for rq in itm[key]:
            reqs += f'{rq}|'
        reqs = reqs[:-1] + "] OR ["
    return reqs[:-5]


def write_ws_headers(ws, ws_headers):
    row = 1
    col = 2
    for key in ws_headers:
        row += 1
        try:
            ws.cell(row=row, column=col).value = key
            ws.cell(row=row, column=col).font = header_font
            ws.cell(row=row, column=col).alignment = header_alignment
            ws.cell(row=row, column=col+1).value = ws_headers[key]
            ws.cell(row=row, column=col+1).font = content_font
        # triggered by the required attributes list
        # and the "allOf" dictionary
        except ValueError:
            if key == "required":
                ws.cell(row=row, column=col).value = key
                ws.cell(row=row, column=col).font = header_font
                ws.cell(row=row, column=col).alignment = header_alignment
                ws.cell(row=row, column=col+1).value = get_required_items(ws_headers[key]) # noqa
                ws.cell(row=row, column=col+1).font = content_font
            # ignore "allOf" dictionary since it is imported
            # inside the "properties" dictionary earlier
            elif key == "allOf":
                pass
            # exception for batch.json schema
            elif key == "anyOf":
                ws.cell(row=row, column=col).value = key
                ws.cell(row=row, column=col).font = header_font
                ws.cell(row=row, column=col).alignment = header_alignment
                ws.cell(row=row, column=col+1).value = get_list_items(ws_headers[key]) # noqa
                ws.cell(row=row, column=col+1).font = content_font
            # exception for adjustment.json schema
            elif key == "oneOf":
                ws.cell(row=row, column=col).value = "one of required"
                ws.cell(row=row, column=col).font = header_font
                ws.cell(row=row, column=col).alignment = header_alignment
                ws.cell(row=row, column=col+1).value = get_contribution_requirements(ws_headers[key]) # noqa
                ws.cell(row=row, column=col+1).font = content_font
            else:
                raise UnhandledException

    return row  # returns the last row populated with header data


def reorder_sheets(wb):
    """
    reorder excel sheets and color tabs appropriately for human readability
    """
    # move enums
    for ws in wb._sheets.copy():
        if "enum" in ws.title:
            tmp = ws
            tmp.sheet_properties.tabColor = ENUM_TAB_COLOR
            wb._sheets.remove(ws)
            wb._sheets.append(tmp)
    # move items
    for ws in wb._sheets.copy():
        if ws.title not in SCHEMA_NAMES and "enum" not in ws.title:
            tmp = ws
            tmp.sheet_properties.tabColor = ITEM_TAB_COLOR
            wb._sheets.remove(ws)
            wb._sheets.append(tmp)
    # color schema tabs
    for ws in wb._sheets:
        if ws.title in SCHEMA_NAMES:
            ws.sheet_properties.tabColor = SCHEMA_TAB_COLOR


def format_excel(wb):
    """
    format excel for human readability
    """
    # remove default sheet
    dflt = [sh for sh in wb.sheetnames if sh.lower().find("sheet") != -1]
    for sh in dflt:
        wb.remove(wb[sh])

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        # list comprehension that returns a list of all the cells
        # in the worksheet which have data validations
        # this effectively returns the list of enum cells
        enum_cells = [ws.cell(cell[0], cell[1]) for validation in ws.data_validations.dataValidation for cell in validation.ranges.ranges[0].cells] # noqa

        dims = {}
        if "enum" in ws_name:
            skip = False
        else:
            skip = True

        # modify cell style and width to fit contents
        for row in ws.rows:
            # for schemas we ignore the metadata at the top of the worksheet
            # since we only want to format the properties and attributes table
            if skip:
                cell = [cl for cl in row if cl.column_letter == "B"][0]
                if cell.value == "property":
                    skip = False
            if not skip:
                for cell in row:
                    if cell.value:
                        # for links to items we only want
                        # the length of the displayed name,
                        # not the whole formula
                        if type(cell.value) == str and "HYPERLINK" in cell.value: # noqa
                            str_val = cell.value.split('"')[-2]
                            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str_val))) # noqa
                        else:
                            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)))) # noqa
                    # beautify cells by greying out blanks
                    # and adding borders to these cells
                    if cell.value is None and cell not in enum_cells and cell.column_letter != "A" and cell.row != 1: # noqa
                        cell.fill = blankFill
                        if "enum" not in ws_name:
                            cell.border = content_border
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    # reorder sheets and colour tabs
    reorder_sheets(wb)


def read_json(path, common_seed=None):
    """
    reads and returns json data for schema
    """
    with open(path) as json_file:
        data = json.load(json_file)
        # entities use the "allOf" attribute
        # to inherit from the <entity common shcema>
        if "allOf" in data:
            entity_data = read_json(ENTITY_SEED, common_seed)
            for prop in data["properties"]:
                entity_data["properties"][prop] = data["properties"][prop]
            data["properties"] = entity_data["properties"]

        # $ref is used to inherit from the <common properties schema>
        if "properties" in data:
            for itm in data["properties"].copy():
                for key in data["properties"][itm].copy():
                    if key == "$ref":
                        common_data = read_json(common_seed)
                        prprty = data["properties"][itm][key].split("/")[-1]
                        for atr in common_data[prprty]:
                            data["properties"][itm][atr] = common_data[prprty][atr] # noqa
                        data["properties"][itm].pop("$ref")
        return data


def jsontoexcel():
    """
    Read Json schemas and write them to the fire_schemas excel
    """
    # Remove excel file if it already exists to ensure only
    # the latest generated file remains after script execution
    if os.path.exists(excel_file):
        os.remove(excel_file)
    wb = Workbook()
    for key in JSON_SCHEMAS:
        json_data = read_json(JSON_SCHEMAS[key], PRODUCT_SEED)
        write_excel(wb, key, json_data)

    format_excel(wb)
    wb.save(excel_file)


def test_jsontoexcel():
    """
    same as jsontoexcel
    doesn't save the workbook
    instead returns it for testing purposes
    """
    wb = Workbook()
    for key in JSON_SCHEMAS:
        json_data = read_json(JSON_SCHEMAS[key], PRODUCT_SEED)
        write_excel(wb, key, json_data)

    format_excel(wb)
    return wb


if __name__ == "__main__":
    jsontoexcel()
    print("Excel file generated succesfully!")
